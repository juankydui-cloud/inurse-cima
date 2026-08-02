#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_nnn_json.py — iNurse
-------------------------------------------------------------------------
Regenera nnn_codes.json a partir de codigos_NNN_curados_iNurse.xlsx.

Uso:
    python3 build_nnn_json.py codigos_NNN_curados_iNurse.xlsx nnn_codes.json

Ejecútalo cada vez que añadas o completes filas en el Excel, y luego llama
a terminology-connector.invalidateNnnCache() en el servidor (o reinícialo)
para que recoja la versión nueva.

Solo se exportan los términos con Estado="Verificado" (los tres códigos
presentes). Los términos a medio rellenar quedan en "pendientes_de_verificar"
— nunca se inventa un código a partir de una fila incompleta.
"""
import sys
import json
import re
import openpyxl


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower() if s else s


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "codigos_NNN_curados_iNurse.xlsx"
    json_path = sys.argv[2] if len(sys.argv) > 2 else "nnn_codes.json"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Códigos NNN"]

    verified, pending = {}, []
    for row in range(3, 28):
        termino = ws.cell(row, 1).value
        if not termino:
            continue

        estado = ws.cell(row, 11).value
        entry = {
            "termino_original": termino,
            "nanda": {"code": ws.cell(row, 2).value, "label": ws.cell(row, 3).value},
            "nic": {"code": ws.cell(row, 4).value, "label": ws.cell(row, 5).value},
            "noc": {"code": ws.cell(row, 6).value, "label": ws.cell(row, 7).value},
            "fuente": ws.cell(row, 8).value,
            "fecha_verificacion": ws.cell(row, 9).value,
            "revisado_por": ws.cell(row, 10).value,
        }

        if estado == "Verificado":
            verified[norm(termino)] = entry
        else:
            pending.append(termino)

    out = {
        "_generado": f"a partir de {xlsx_path} — regenerar con build_nnn_json.py",
        "_aviso": "Solo se incluyen términos con Estado=Verificado (códigos confirmados manualmente en NNNConsult). Nada inventado.",
        "verificados": verified,
        "pendientes_de_verificar": pending,
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"✓ {len(verified)} términos verificados · {len(pending)} pendientes → {json_path}")


if __name__ == "__main__":
    main()
