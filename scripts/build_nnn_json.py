#!/usr/bin/env python3
"""Regenera data/nnn_codes.json a partir de codigos_NNN_curados_iNurse.xlsx.

Solo se incluyen filas con Estado == "Verificado" (NANDA + NIC + NOC
rellenados). Las filas pendientes se ignoran: nunca se fabrica un código.

Uso:
    pip install openpyxl
    python3 scripts/build_nnn_json.py [ruta_al_xlsx]
"""
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "data" / "codigos_NNN_curados_iNurse.xlsx"
OUTPUT_JSON = ROOT / "data" / "nnn_codes.json"
SHEET_NAME = "Códigos NNN"


def build(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"La pestaña '{SHEET_NAME}' está vacía.")

    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {name: header.index(name) for name in ("Término", "NANDA", "NIC", "NOC", "Estado", "Verificado por", "Fecha verificación") if name in header}

    terms = []
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        estado = row[idx["Estado"]] if "Estado" in idx else None
        if str(estado).strip() != "Verificado":
            continue

        termino = row[idx["Término"]]
        nanda = row[idx["NANDA"]]
        nic = row[idx["NIC"]]
        noc = row[idx["NOC"]]
        if not (termino and nanda and nic and noc):
            continue

        entry = {
            "termino": str(termino).strip(),
            "nanda": str(nanda).strip(),
            "nic": str(nic).strip(),
            "noc": str(noc).strip(),
        }
        if "Verificado por" in idx and row[idx["Verificado por"]]:
            entry["verifiedBy"] = str(row[idx["Verificado por"]]).strip()
        if "Fecha verificación" in idx and row[idx["Fecha verificación"]]:
            fecha = row[idx["Fecha verificación"]]
            entry["verifiedAt"] = fecha.isoformat() if hasattr(fecha, "isoformat") else str(fecha)
        terms.append(entry)

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "generatedBy": "scripts/build_nnn_json.py",
        "source": f"{xlsx_path.name} (pestaña {SHEET_NAME}, filas con Estado=Verificado)",
        "terms": terms,
    }


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"No se encuentra el Excel: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    data = build(xlsx_path)
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"{len(data['terms'])} término(s) verificado(s) escritos en {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
